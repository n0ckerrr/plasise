import requests
import re
import os
import sys
import shutil
import mysql.connector
from pathlib import Path
from datetime import datetime
import openpyxl
from dotenv import load_dotenv

# Load environment variables
env_path = os.path.join(os.path.dirname(__file__), '..', '.env')
if not os.path.exists(env_path):
    env_path = "/code/.env"
load_dotenv(env_path)

# Configuration
LOGIN_URL = "https://www.ibdglobal.com/web/login"
DOWNLOAD_URL = "https://www.ibdglobal.com/my/pricelist/download/simple"
USERNAME = os.getenv("IBD_USER")
PASSWORD = os.getenv("IBD_PASS")

# Database Configuration (from env)
DB_HOST = os.getenv('DB_HOST', '127.0.0.1')
DB_PORT = int(os.getenv('DB_PORT', 9966))
DB_USER = os.getenv('DB_USER', 'plasise')
DB_PASSWORD = os.getenv('DB_PASSWORD', '')
DB_NAME = os.getenv('DB_NAME', 'plasise')

# Paths
if os.name == 'nt':
    OUTPUT_DIR = Path("c:/Users/plasi/Documents/PROYECTOS/PROYECTOS-ANTIGRAVITY/.tmp")
else:
    OUTPUT_DIR = Path("/etc/easypanel/projects/n0cker/plasise/code/.tmp")

DOWNLOAD_FILE = OUTPUT_DIR / "ibd_productos.csv"
EXCEL_FILE = OUTPUT_DIR / "ibd_productos.xlsx"

def get_db_connection():
    try:
        return mysql.connector.connect(
            host=DB_HOST,
            port=DB_PORT,
            user=DB_USER,
            password=DB_PASSWORD,
            database=DB_NAME
        )
    except Exception as e:
        print(f"[!] Error connecting to database: {e}")
        return None

def log_event(event_type, message):
    try:
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO script_events (script_name, event_type, message) VALUES (%s, %s, %s)",
                ('sync_ibd_prices', event_type, str(message)[:1000])
            )
            conn.commit()
            cursor.close()
            conn.close()
    except Exception as e:
        print(f"[!] Unable to log event: {e}")

def download_ibd_file():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    session = requests.Session()
    
    # 1. Login Page
    print(f"[*] Accessing login page: {LOGIN_URL}")
    try:
        response = session.get(LOGIN_URL)
    except Exception as e:
        print(f"[!] Request failed: {e}")
        return False

    csrf_token_match = re.search(r'name="csrf_token" value="([^"]+)"', response.text)
    if not csrf_token_match:
        print("[!] Error: No csrf_token found")
        return False
    token = csrf_token_match.group(1)
    
    # 2. Login
    login_data = {
        'login': USERNAME,
        'password': PASSWORD,
        'csrf_token': token,
        'redirect': ''
    }
    print("[*] Logging in...")
    response = session.post(LOGIN_URL, data=login_data)
    
    if response.status_code != 200 or "error" in response.text.lower():
        print("[!] Login failed or error in response.")
        return False
        
    # 3. Download
    print(f"[*] Downloading file from: {DOWNLOAD_URL}")
    response = session.get(DOWNLOAD_URL, stream=True)
    
    if response.status_code == 200:
        if "login" in response.url:
            print("[!] Error: Redirected to login page during download.")
            return False
            
        with open(DOWNLOAD_FILE, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                f.write(chunk)
        print(f"[+] Download successful: {DOWNLOAD_FILE} ({os.path.getsize(DOWNLOAD_FILE)} bytes)")
        
        # Rename to xlsx
        try:
            shutil.copy(DOWNLOAD_FILE, EXCEL_FILE)
            print(f"[+] Renamed to {EXCEL_FILE}")
            return True
        except Exception as e:
            print(f"[!] Error renaming file: {e}")
            return False
    else:
        print(f"[!] Download failed with status: {response.status_code}")
        return False

def sync_prices():
    log_event('INFO', 'Iniciando proceso de sincronización de precios IBD Global')
    if not download_ibd_file():
        log_event('ERROR', 'Fallo al descargar el archivo de precios de IBD.')
        return
    
    print("[*] Starting database sync...")
    conn = get_db_connection()
    if not conn:
        return
    
    cursor = conn.cursor(dictionary=True)
    
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
        sheet = wb.active
        
        # Get Headers
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        # Map headers to indices
        try:
            idx_sku = headers.index('Code')
            idx_price = headers.index('Price')
            idx_msrp = headers.index('MSRP')
            idx_stock = headers.index('Stock')
        except ValueError as e:
            print(f"[!] Header missing: {e}")
            print(f"Headers found: {headers}")
            return

        updated_count = 0
        total_count = 0
        
        # Iterate Rowsss
        for row in sheet.iter_rows(min_row=2, values_only=True):
            sku = row[idx_sku]
            price = row[idx_price] # Coste
            msrp = row[idx_msrp]   # PVP Recomendado
            stock = row[idx_stock]
            
            if not sku:
                continue
            
            total_count += 1
            
            # Find product in DB
            cursor.execute("SELECT id, coste, stock_actual, precio_pvp_recomendado FROM productos WHERE sku = %s", (sku,))
            product = cursor.fetchone()
            
            if product:
                updates = []
                values = []
                
                # Check Coste
                if price is not None:
                    try:
                        price_float = float(price)
                        # Only update if difference is significant
                        if product['coste'] is None or abs(float(product['coste']) - price_float) > 0.01:
                            updates.append("coste = %s")
                            values.append(price_float)
                    except:
                        pass
                
                # Check SRP
                if msrp is not None:
                    try:
                        msrp_float = float(msrp)
                        if product['precio_pvp_recomendado'] is None or abs(float(product['precio_pvp_recomendado']) - msrp_float) > 0.01:
                            updates.append("precio_pvp_recomendado = %s")
                            values.append(msrp_float)
                    except:
                        pass

                # Check Stock
                if stock is not None:
                    try:
                        # Normalize stock (ibd sometimes uses >100 strings?)
                        # Assuming integer for now based on previous inspection (it was 0)
                        stock_int = int(stock)
                        if product['stock_actual'] is None or int(product['stock_actual']) != stock_int:
                            updates.append("stock_actual = %s")
                            values.append(stock_int)
                    except:
                        pass
                
                if updates:
                    values.append(product['id'])
                    query = f"UPDATE productos SET {', '.join(updates)}, fecha_actualizacion = NOW() WHERE id = %s"
                    cursor.execute(query, values)
                    updated_count += 1
                    # print(f"Updated {sku}: {updates}") # Verbose

        conn.commit()
        msg = f"Sincronización finalizada. {total_count} filas procesadas. {updated_count} productos actualizados."
        print(f"[SUCCESS] {msg}")
        log_event('SUCCESS', msg)
        
    except Exception as e:
        print(f"[!] Error during sync: {e}")
        log_event('ERROR', f'Error durante la sincronización: {e}')
    finally:
        if 'wb' in locals():
            wb.close()
        cursor.close()
        conn.close()

if __name__ == "__main__":
    sync_prices()
